import openpyxl
from openpyxl.utils import get_column_letter
from decimal import Decimal
from urllib.request import urlopen
from xml.etree import ElementTree


def update_ecb_reference_rates():
    ecb_data = get_eurofxref_daily()
    for row_d in ecb_data:
        currency_date = row_d.attrib["time"]
        try:
            workbook = openpyxl.load_workbook('ecb_reference_rates.xlsx')
            workbook.sheetnames[-1]
            if workbook.sheetnames:
                last_sheet_date = workbook.sheetnames[-1]
                if currency_date != last_sheet_date:                    
                    sheet = create_sheet(workbook.create_sheet(), currency_date, row_d)
                    adjust_column_width_from_col(sheet)
                    workbook.move_sheet(currency_date, -(len(workbook.sheetnames)-1))
        except Exception:
            workbook = openpyxl.Workbook()
            sheet = create_sheet(workbook.active, currency_date, row_d)
            adjust_column_width_from_col(sheet)            

    workbook.save('ecb_reference_rates.xlsx')

def get_eurofxref_daily():
    file = urlopen('https://www.ecb.europa.eu/stats/eurofxref/eurofxref-daily.xml')
    data_file = file.read()
    file.close()
    envelope = ElementTree.fromstring(data_file)
    namespaces = {
        "gesmes": "http://www.gesmes.org/xml/2002-08-01",
        "eurofxref": "http://www.ecb.int/vocabulary/2002-08-01/eurofxref",
    }

    return envelope.findall("./eurofxref:Cube/eurofxref:Cube[@time]", namespaces)

def create_sheet(sheet, date, row_d):
    titles = ['CURRENCY', 'RATE', 'DATE']
    sheet.append(titles)
    for cr in list(row_d):
        row = [cr.attrib["currency"], Decimal(cr.attrib["rate"]), date]
        sheet.title = date
        sheet.append(row)
    return sheet

def adjust_column_width_from_col(sheet):
    sheet.auto_filter.ref = 'A1:B32'
    sheet.auto_filter.add_filter_column(1, ['USD', 'BGN'])
    sheet.auto_filter.add_sort_condition('B2:B32')    
    for i in range(1, sheet.max_column+1):
        sheet.column_dimensions[get_column_letter(i)].bestFit = True
        sheet.column_dimensions[get_column_letter(i)].auto_size = True

def get_currency_rate_by_code(code):
    try:
        workbook = openpyxl.load_workbook('ecb_reference_rates.xlsx')
        if workbook.sheetnames:
            last_sheet = workbook.active
            for row in last_sheet.iter_rows():
                if row[0].value == code:
                    currency_rate = row[1].value
                    return currency_rate
    except Exception:
        pass

def main():
    update_ecb_reference_rates()
    currency_code = 'USD'
    currency_rate = get_currency_rate_by_code(currency_code)
    print(currency_rate)

if __name__ == "__main__":
    main()
